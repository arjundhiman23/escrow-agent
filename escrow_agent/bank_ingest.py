"""Ingest Axis CATRA bank-format statements and classify to ATSL categories.

Bank format columns: FORACID, ACCT_NAME, ACCT_OPN_DATE, CUST_ID, File name,
SCHM_CODE, SCHM_TYPE, TRAN_ID, TRAN_DATE, VALUE_DATE, TRAN_TYPE, PART_TRAN_TYPE,
TRAN_SUB_TYPE, TRAN_PARTICULAR, TRAN_AMT, BALANCE, PART_TRAN_SRL_NUM, Remarks.

ATSL categories (per bank's ATSL output format):
  Credits: Proceeds from Equity | From Redemption of Investments | Other Revenue |
           Internal Company transfer | Other Refunds | Proceed from Term Loan
  Debits:  O&M Exp/Project Construction Payments | Debt servicing | Reserve creations |
           Permitted Investments | Statutory Payments | Surplus distribution to Borrower
"""
from dataclasses import dataclass, field
from openpyxl import load_workbook

CREDIT_CATS = ["Proceeds from Equity", "From Redemption of Investments", "Other Revenue",
               "Internal Company transfer", "Other Refunds", "Proceed from Term Loan"]
DEBIT_CATS = ["O&M Exp/Project Construction Payments", "Debt servicing", "Reserve creations",
              "Permitted Investments", "Statutory Payments", "Surplus distribution to Borrower"]


@dataclass
class Txn:
    quarter: str
    date: object
    dc: str                 # 'C' / 'D'
    narration: str
    amount: float
    balance: float
    bank_remark: str        # bank's own label from Remarks col (may be None)
    category: str = ""
    basis: str = ""         # rule that fired
    review: bool = False
    conflict: str = ""
    ai_suggestion: str = ""
    ai_confidence: float = 0.0


def read_quarter(path, quarter, account):
    wb = load_workbook(path, read_only=True)
    ws = wb["Sheet1"]
    txns = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]) != str(account):
            continue
        row = tuple(row) + (None,) * (18 - len(row))
        amt = row[14]
        if amt is None:      # blank/summary rows in bank export
            continue
        txns.append(Txn(quarter=quarter, date=row[8], dc=str(row[11]).strip().upper(),
                        narration=" ".join(str(row[13] or "").split()),
                        amount=float(amt), balance=float(row[15]) if row[15] is not None else None,
                        bank_remark=(str(row[17]).strip() if row[17] not in (None, "None") else "")))
    wb.close()
    return txns


def classify(t: Txn, loan_accounts=("923060049840106", "923060049840119"), related_entities=()):
    n = t.narration.lower()
    if t.dc == "C":
        if "disbursement credit" in n or any(a in t.narration for a in loan_accounts):
            return "Proceed from Term Loan", "narration: loan disbursement"
        # known related/group entity as counterparty -> internal company transfer, even when the
        # narration is just the counterparty name with no "transfer" keyword (matches bank convention
        # of classifying by who sent the money, not just what the narration literally says)
        if any(e.lower() in n for e in related_entities if e):
            return "Internal Company transfer", "narration: known related entity counterparty"
        if any(k in n for k in ("mutual fund", "mutua", "fd matur", "fd closure", "trd", "fixed deposit",
                                "flexi deposit", "deposit prin")):
            return "From Redemption of Investments", "narration: investment redemption"
        if any(k in n for k in ("inter company transfer", "intercompany transfer", "inter-company transfer",
                                "internal company transfer")):
            return "Internal Company transfer", "narration: inter-company transfer"
        # explicit refund narrations (tax refunds, balance refunds, GST refunds, vendor refunds, etc.)
        # map to the bank's dedicated "Other Refunds" ATSL category -> distinct from generic Other Revenue.
        # Checked before the expense-narration fallback below since "refund" is a more specific, confident
        # signal than the ambiguous expense-reversal guess.
        if "refund" in n:
            return "Other Refunds", "narration: refund"
        # credit bearing an expense narration = probable reversal/refund. Bank convention
        # (note * on ATSL sheet) routes all non-redemption, non-loan credits to Other
        # Revenue; we follow it for reconciliation but flag the row as an observation.
        if any(k in n for k in ("o&m", "o and m", "epc", "statutory", "debt servic")):
            return "Other Revenue", "credit with expense narration — probable reversal/refund (flag for review)"
        return "Other Revenue", "default credit (NHAI/sponsor/other receipts; DOCC not achieved)"
    else:
        if any(k in n for k in ("statutory", "gst", "tds", "advance tax", "income tax", "stat pay")):
            return "Statutory Payments", "narration: statutory"
        if any(k in n for k in ("debt servic", "debt service", "cash sweep")):
            return "Debt servicing", "narration: debt servicing"
        if any(k in n for k in ("liquid fund", "mutual fund", "fixed deposit", " fd ", "investment",
                                "term deposit")) or n.startswith("td/"):
            return "Permitted Investments", "narration: investment placement"
        if any(k in n for k in ("dsra", "mmr", "reserve")):
            return "Reserve creations", "narration: reserve"
        if any(k in n for k in ("dividend", "surplus", "distribution")):
            return "Surplus distribution to Borrower", "narration: distribution"
        if any(k in n for k in ("o&m", "o and m", "epc", "construction", "capex")):
            return "O&M Exp/Project Construction Payments", "narration: O&M/EPC"
        return "O&M Exp/Project Construction Payments", "default debit -> O&M/Construction (flag for review)"


# bank Remarks label -> ATSL category (bank remarks keep statutory inside O&M; ATSL splits it)
REMARK_MAP = {
    "o&m exp /project construction payments": "O&M Exp/Project Construction Payments",
    "debt servicing^": "Debt servicing",
    "permitted investments": "Permitted Investments",
    "from redemption of investments": "From Redemption of Investments",
    "proceed from term loan": "Proceed from Term Loan",
    "reserve creations": "Reserve creations",
    "surplus distribution to borrower": "Surplus distribution to Borrower",
}


def classify_all(txns, ai_assist=False, ai_model="claude-haiku-4-5-20251001", tracker=None, related_entities=()):
    """Rule-classify every transaction, then (optionally) run a cheap AI pass
    over just the ones flagged for review — never overrides a rule match.
    `related_entities` is a deal-specific, admin-editable list of known group/
    affiliate company names (from the deal profile) so credits from sister
    SPVs are recognised as Internal Company Transfer even without a "transfer"
    keyword in the narration.
    """
    for t in txns:
        t.category, t.basis = classify(t, related_entities=related_entities)
        if t.basis.endswith("(flag for review)"):
            t.review = True
        # conflict check vs bank's own remark where present
        if t.bank_remark:
            mapped = REMARK_MAP.get(" ".join(t.bank_remark.lower().replace("\xa0", " ").split()))
            if mapped and mapped != t.category:
                # statutory narrations are the known, intended divergence (ATSL splits them out)
                if not (t.category == "Statutory Payments"
                        and mapped == "O&M Exp/Project Construction Payments"):
                    t.conflict = f"bank remark '{t.bank_remark}' -> {mapped}, rules -> {t.category}"
                    t.review = True

    if ai_assist:
        pending = [t for t in txns if t.review or t.conflict]
        if pending:
            _ai_assist_classify(pending, ai_model, tracker)
    return txns


def _ai_assist_classify(pending, model, tracker):
    """Second opinion from an LLM on rule-flagged rows only. Suggestions are
    recorded as `ai_suggestion` for the analyst to accept — they do NOT
    silently change `category`, so the ATSL totals stay rule-based and
    reproducible; this only reduces manual review effort.
    """
    try:
        import anthropic, json
    except ImportError:
        return
    client = anthropic.Anthropic()
    all_cats = [f"{c} (credit)" for c in CREDIT_CATS] + [f"{c} (debit)" for c in DEBIT_CATS]
    cats_text = "\n".join(f"- {c}" for c in all_cats)
    batch_size = 25
    for start in range(0, len(pending), batch_size):
        batch = pending[start:start + batch_size]
        items = "\n".join(
            f'{i}. {"credit" if t.dc=="C" else "debit"} amount={t.amount:.2f} narration="{t.narration}"'
            for i, t in enumerate(batch))
        prompt = (f"Classify these escrow account bank transactions into ATSL categories.\n\n"
                 f"Categories:\n{cats_text}\n\nTransactions:\n{items}\n\n"
                 'Respond with ONLY a JSON array: [{"i": 0, "category": "<exact category name>", "confidence": 0.0-1.0}]')
        try:
            resp = client.messages.create(model=model, max_tokens=40 * len(batch) + 100,
                                          messages=[{"role": "user", "content": prompt}])
            if tracker is not None:
                tracker.record("classification", model, resp.usage)
            text = "".join(b.text for b in resp.content if b.type == "text")
            text = text.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            results = json.loads(text)
            for r in results:
                t = batch[int(r["i"])]
                cat = r.get("category", "")
                if cat in CREDIT_CATS or cat in DEBIT_CATS:
                    t.ai_suggestion = cat
                    t.ai_confidence = float(r.get("confidence", 0))
        except Exception:
            continue


def quarter_summary(txns):
    """Per-quarter category totals + opening/closing balances."""
    out = {}
    for t in txns:
        q = out.setdefault(t.quarter, {"credits": {c: 0.0 for c in CREDIT_CATS},
                                       "debits": {c: 0.0 for c in DEBIT_CATS},
                                       "first": None, "last": None})
        (q["credits"] if t.dc == "C" else q["debits"])[t.category] += t.amount
        if t.balance is not None and t.date is not None:
            if q["first"] is None or t.date < q["first"][0]:
                q["first"] = (t.date, t.balance, t.amount, t.dc)
            if q["last"] is None or t.date >= q["last"][0]:
                q["last"] = (t.date, t.balance, t.amount, t.dc)
    for q in out.values():
        q["total_credit"] = sum(q["credits"].values())
        q["total_debit"] = sum(q["debits"].values())
        # closing = balance on last dated row; opening derived: closing - net movement
        q["closing"] = q["last"][1] if q["last"] else None
        if q["closing"] is not None:
            q["opening"] = round(q["closing"] - q["total_credit"] + q["total_debit"], 2)
    return out
