"""Build the CATRA master file from the KB structured extract.

CATRA = classification framework for the escrow (main) account:
  - debit-side categories = Escrow Agreement Order of Priority (waterfall priorities)
  - credit-side categories = Escrow Agreement permitted deposits
  - keyword rules for classifying bank narrations into these categories
Outputs:
  - CATRA_Master.xlsx (analyst-facing master)
  - config/categories.yaml regenerated so the transaction pipeline classifies
    against the deal-derived taxonomy (admin-editable, no code deployment).
"""
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# Deal-specific keyword rules per category (admin-editable in categories.yaml after generation)
DEBIT_KEYWORDS = {
    "TAXES":      ["gst", "tds", "income tax", "advance tax", "cess", "duty", "levy", "statutory"],
    "CONSTR":     ["construction", "epc", "contractor bill", "capex", "mobilisation", "mobilization", "ra bill"],
    "OM_EXP":     ["o&m", "o & m", "operation and maintenance", "routine maintenance", "opex"],
    "NHAI_OM":    ["nhai o&m", "authority o&m", "nhai maintenance"],
    "NHAI_PAY":   ["nhai payment", "payable to nhai", "authority payment", "concession fee"],
    "DEBT_SERV":  ["interest", "principal", "repayment", "instalment", "installment", "emi", "debt service", "loan servicing"],
    "NHAI_DMG":   ["damages", "penalty to nhai", "liquidated damages", "ld payment"],
    "SUB_DEBT":   ["subordinate debt", "sub debt", "sub-debt"],
    "WCR":        ["working capital reserve", "wcr"],
    "DSRA":       ["dsra", "debt service reserve"],
    "MMRA":       ["mmr", "major maintenance"],
    "ADD_OM":     ["additional o&m", "excess o&m"],
    "CASH_SWEEP": ["cash sweep", "prepayment", "pre-payment", "pre payment"],
    "DISTRIB":    ["dividend", "distribution", "restricted payment", "surplus distribution"],
}
CREDIT_KEYWORDS = {
    "EQUITY":     ["equity", "share subscription", "share capital", "share application"],
    "DISBURSE":   ["disbursement", "loan disbursal", "facility drawdown", "drawdown"],
    "ANNUITY":    ["annuity", "nhai annuity", "bonus", "o&m receipt", "interest receipt"],
    "PROCEEDS":   ["mobilization advance", "mobilisation advance", "insurance claim", "insurance proceeds",
                   "grant", "termination payment", "promoter contribution", "sponsor contribution"],
    "RETENTION":  ["retention account", "sub account transfer", "sub-account transfer"],
    "INV_INCOME": ["fd interest", "fixed deposit interest", "investment income", "fd maturity", "mutual fund redemption"],
    "OTHER_DEP":  ["misc receipt", "other receipt"],
}


def load_extract(path="kb/structured/deal_extract.yaml"):
    with open(path) as f:
        return yaml.safe_load(f)


def build_categories_yaml(extract, out_path="config/categories.yaml"):
    cats = []
    for row in extract["order_of_priority"]:
        cats.append({
            "code": row["code"], "name": row["name"], "side": "debit",
            "waterfall_priority": row["priority"],
            "agreement_item": f"2.3(B)({row['item']})",
            "sub_account": row["sub_account"],
            "keywords": DEBIT_KEYWORDS.get(row["code"], []),
            "description": row["desc"],
        })
    for row in extract["permitted_deposits"]:
        cats.append({
            "code": row["code"], "name": row["name"], "side": "credit",
            "agreement_item": f"2.3(A)({row['item']})",
            "keywords": CREDIT_KEYWORDS.get(row["code"], []),
        })
    doc = {
        "taxonomy_source": "Generated from KB: Escrow Agreement cl. 2.3 + Sanction Letter cl. 20 (KLEPL / Axis Bank). Admin-editable.",
        "categories": cats,
        "internal_transfer_keywords": ["retention account", "sub account", "sub-account", "own account transfer", "internal transfer"],
    }
    with open(out_path, "w") as f:
        yaml.safe_dump(doc, f, sort_keys=False, width=110)
    return len(cats)


def _hdr(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws, r, c, v, bold=False, alt=False, wrap=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name=ARIAL, size=10, bold=bold)
    cell.border = THIN
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if alt:
        cell.fill = ALT_FILL
    return cell


def build_catra_xlsx(extract, out_path="output/CATRA_Master.xlsx"):
    wb = Workbook()

    ws = wb.active
    ws.title = "Debit Categories (Waterfall)"
    _hdr(ws, 1, ["Priority", "Code", "Category", "Sub-Account", "Agreement Ref", "Sanction Ref (cl.20)",
                 "Keywords", "Description"], [9, 12, 26, 30, 14, 18, 40, 55])
    san_by_code = {}
    for s in extract["sanction_waterfall"]:
        san_by_code.setdefault(s["maps_to"], []).append(str(s["step"]))
    for i, row in enumerate(extract["order_of_priority"]):
        r, alt = i + 2, i % 2 == 1
        _cell(ws, r, 1, row["priority"], alt=alt)
        _cell(ws, r, 2, row["code"], bold=True, alt=alt)
        _cell(ws, r, 3, row["name"], alt=alt)
        _cell(ws, r, 4, row["sub_account"], alt=alt)
        _cell(ws, r, 5, f"2.3(B)({row['item']})", alt=alt)
        _cell(ws, r, 6, "step " + ", ".join(san_by_code.get(row["code"], ["—"])), alt=alt)
        _cell(ws, r, 7, ", ".join(DEBIT_KEYWORDS.get(row["code"], [])), alt=alt)
        _cell(ws, r, 8, row["desc"], alt=alt)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Credit Categories (Deposits)")
    _hdr(ws2, 1, ["Code", "Permitted Deposit", "Agreement Ref", "Keywords"], [14, 60, 14, 55])
    for i, row in enumerate(extract["permitted_deposits"]):
        r, alt = i + 2, i % 2 == 1
        _cell(ws2, r, 1, row["code"], bold=True, alt=alt)
        _cell(ws2, r, 2, row["name"], alt=alt)
        _cell(ws2, r, 3, f"2.3(A)({row['item']})", alt=alt)
        _cell(ws2, r, 4, ", ".join(CREDIT_KEYWORDS.get(row["code"], [])), alt=alt)
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Sanction Waterfall Map")
    _hdr(ws3, 1, ["Sanction Step", "Sanction Text (cl.20)", "Maps to CATRA Code"], [12, 70, 20])
    for i, s in enumerate(extract["sanction_waterfall"]):
        r, alt = i + 2, i % 2 == 1
        _cell(ws3, r, 1, s["step"], alt=alt)
        _cell(ws3, r, 2, s["text"], alt=alt)
        _cell(ws3, r, 3, s["maps_to"], bold=True, alt=alt)
    rr = len(extract["sanction_waterfall"]) + 3
    _cell(ws3, rr, 1, "Divergences", bold=True, wrap=False)
    for j, d in enumerate(extract["waterfall_divergences"]):
        _cell(ws3, rr + 1 + j, 2, d)

    ws4 = wb.create_sheet("Covenants")
    _hdr(ws4, 1, ["ID", "Source", "Rule"], [8, 28, 100])
    for i, cv in enumerate(extract["covenants"]):
        r, alt = i + 2, i % 2 == 1
        _cell(ws4, r, 1, cv["id"], bold=True, alt=alt)
        _cell(ws4, r, 2, cv["source"], alt=alt)
        _cell(ws4, r, 3, cv["rule"], alt=alt)
    ws4.freeze_panes = "A2"

    wb.save(out_path)
    return out_path
