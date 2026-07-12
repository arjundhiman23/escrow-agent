"""Excel report generation (BRD §6.6, §8.2).

Workbook sheets:
  Transactions        every parsed transaction with classification & flags
  Quarter Summary     category × quarter grid with SUM formulas (Q1..Q4 + FY)
  Variance Analysis   actual vs projected per category per quarter, formulas
  QoQ Movement        quarter-on-quarter deltas (actuals & projections)
  Exceptions Log      waterfall/condition violations + manual-check conditions
  Review Queue        analyst review queue + escalated conflicts (§6.2)
  Audit Trail         timestamped processing log (§7)

Totals and variances are written as Excel FORMULAS so the workbook
recalculates if an analyst adjusts a figure.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Taxonomy, Transaction

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, bold=True, size=10)
SECTION_FONT = Font(name=FONT, bold=True, size=11, color="1F4E79")
FLAG_FILL = PatternFill("solid", fgColor="FCE4EC")
MATERIAL_FILL = PatternFill("solid", fgColor="FFF3CD")
THIN = Border(*[Side(style="thin", color="B0B0B0")] * 4)
AMT_FMT = "#,##0.00;(#,##0.00);-"
PCT_FMT = "0.0%"
QUARTERS = ("Q1", "Q2", "Q3", "Q4")


def _header(ws, row, headers, widths=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _cell(ws, row, col, value, fmt=None, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = BOLD_FONT if bold else BASE_FONT
    c.border = THIN
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def generate(out_path: str | Path, txns: list[Transaction], taxonomy: Taxonomy,
             agg: dict, var: dict, wf_exceptions: list[dict],
             manual_conds: list[dict], sanction: dict, audit) -> Path:
    wb = Workbook()

    _sheet_transactions(wb.active, txns)
    _sheet_quarter_summary(wb.create_sheet("Quarter Summary"), taxonomy, agg)
    _sheet_variance(wb.create_sheet("Variance Analysis"), var)
    _sheet_qoq(wb.create_sheet("QoQ Movement"), var)
    _sheet_exceptions(wb.create_sheet("Exceptions Log"), wf_exceptions, manual_conds)
    _sheet_review_queue(wb.create_sheet("Review Queue"), txns)
    _sheet_audit(wb.create_sheet("Audit Trail"), audit)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    audit.log("REPORT_GENERATION", f"Excel workbook written: {out_path.name}")
    _ = sanction
    return out_path


# ---------------------------------------------------------------------------
def _sheet_transactions(ws, txns):
    ws.title = "Transactions"
    headers = ["Row", "Date", "Quarter", "Type", "Amount (INR)", "Narration",
               "Balance (INR)", "Category Code", "Category", "Source",
               "Confidence", "Internal Transfer", "Duplicate", "Balance Break",
               "Waterfall Violation", "Review / Conflict"]
    _header(ws, 1, headers, [6, 12, 8, 6, 16, 48, 16, 14, 30, 8, 11, 14, 10, 12, 24, 34])
    r = 2
    for t in txns:
        flagged = t.is_duplicate or t.balance_break or bool(t.waterfall_violation) \
                  or t.needs_review or t.is_escalated_conflict
        fill = FLAG_FILL if flagged else None
        _cell(ws, r, 1, t.row_no, fill=fill)
        _cell(ws, r, 2, t.txn_date.strftime("%d-%m-%Y"), fill=fill)
        _cell(ws, r, 3, t.quarter, fill=fill)
        _cell(ws, r, 4, t.txn_type, fill=fill)
        _cell(ws, r, 5, t.amount, AMT_FMT, fill=fill)
        _cell(ws, r, 6, t.narration, fill=fill)
        _cell(ws, r, 7, t.balance, AMT_FMT, fill=fill)
        _cell(ws, r, 8, t.category_code or "—", fill=fill)
        _cell(ws, r, 9, t.category_name or "UNCLASSIFIED", fill=fill)
        _cell(ws, r, 10, t.classification_source, fill=fill)
        _cell(ws, r, 11, t.confidence if t.category_code else None, "0.00", fill=fill)
        _cell(ws, r, 12, "YES" if t.is_internal_transfer else "", fill=fill)
        _cell(ws, r, 13, "DUPLICATE" if t.is_duplicate else "", fill=fill)
        _cell(ws, r, 14, "BREAK" if t.balance_break else "", fill=fill)
        _cell(ws, r, 15, t.waterfall_violation, fill=fill)
        _cell(ws, r, 16, t.review_reason or t.conflict_reason, fill=fill)
        r += 1


def _sheet_quarter_summary(ws, taxonomy, agg):
    """Category × quarter grid. Category cells are values (computed by the
    engine); totals/FY columns are SUM formulas so the sheet recalculates."""
    _header(ws, 1, ["Category (per TRA/CATRA taxonomy)", "Q1", "Q2", "Q3", "Q4",
                    "Full Year"], [46, 16, 16, 16, 16, 18])
    r = 2

    def block(title, cats, side):
        nonlocal r
        c = ws.cell(row=r, column=1, value=title)
        c.font = SECTION_FONT
        r += 1
        first = r
        for cat in cats:
            _cell(ws, r, 1, f"{cat.name} [{cat.code}]")
            for qi, q in enumerate(QUARTERS, start=2):
                _cell(ws, r, qi, agg["by_category"].get(q, {}).get(cat.code, 0.0), AMT_FMT)
            _cell(ws, r, 6, f"=SUM(B{r}:E{r})", AMT_FMT)
            r += 1
        # Unclassified line so column totals reconcile to bank totals
        _cell(ws, r, 1, "Unclassified (pending review)")
        for qi, q in enumerate(QUARTERS, start=2):
            key = "Cr" if side == "inflow" else "Dr"
            _cell(ws, r, qi, agg["unclassified"].get(q, {}).get(key, 0.0), AMT_FMT)
        _cell(ws, r, 6, f"=SUM(B{r}:E{r})", AMT_FMT)
        r += 1
        _cell(ws, r, 1, f"Total {title}", bold=True)
        for col in range(2, 7):
            L = get_column_letter(col)
            _cell(ws, r, col, f"=SUM({L}{first}:{L}{r - 1})", AMT_FMT, bold=True)
        total_row = r
        r += 2
        return total_row

    cr_total = block("Credits / Inflows", taxonomy.inflows, "inflow")
    dr_total = block("Debits / Outflows", taxonomy.outflows, "outflow")

    _cell(ws, r, 1, "Net Movement (Inflows − Outflows)", bold=True)
    for col in range(2, 7):
        L = get_column_letter(col)
        _cell(ws, r, col, f"={L}{cr_total}-{L}{dr_total}", AMT_FMT, bold=True)
    r += 2

    c = ws.cell(row=r, column=1, value="Of which: Internal Company Transfers (flagged separately per BRD §5.4)")
    c.font = SECTION_FONT
    r += 1
    for label, key in (("Internal transfers — Credits", "Cr"),
                       ("Internal transfers — Debits", "Dr")):
        _cell(ws, r, 1, label)
        for qi, q in enumerate(QUARTERS, start=2):
            _cell(ws, r, qi, agg["internal_transfers"].get(q, {}).get(key, 0.0), AMT_FMT)
        _cell(ws, r, 6, f"=SUM(B{r}:E{r})", AMT_FMT)
        r += 1
    r += 1
    note = ws.cell(row=r, column=1,
                   value="Note: internal transfer amounts are a memo view of rows flagged as own-account "
                         "transfers; they are included in the credit/debit totals above. Duplicates are excluded.")
    note.font = Font(name=FONT, italic=True, size=9)


def _sheet_variance(ws, var):
    headers = ["Quarter", "Side", "Category", "Actual (INR)", "Projected (INR)",
               "Variance (INR)", "Variance %", "Material?", "Remarks"]
    _header(ws, 1, headers, [9, 9, 40, 17, 17, 17, 11, 10, 60])
    r = 2
    for row in var["rows"]:
        fill = MATERIAL_FILL if row["material"] else None
        _cell(ws, r, 1, row["quarter"], fill=fill)
        _cell(ws, r, 2, row["side"], fill=fill)
        _cell(ws, r, 3, f'{row["name"]} [{row["code"]}]', fill=fill)
        _cell(ws, r, 4, row["actual"], AMT_FMT, fill=fill)
        _cell(ws, r, 5, row["projected"], AMT_FMT, fill=fill)
        _cell(ws, r, 6, f"=D{r}-E{r}", AMT_FMT, fill=fill)
        _cell(ws, r, 7, f'=IF(E{r}=0,"n/a",(D{r}-E{r})/E{r})', PCT_FMT, fill=fill)
        _cell(ws, r, 8, "MATERIAL" if row["material"] else "", bold=row["material"], fill=fill)
        _cell(ws, r, 9, row["remarks"], fill=fill)
        r += 1
    r += 1
    c = ws.cell(row=r, column=1, value="Consolidated Exceptions Summary (most material deviations per quarter — BRD §6.5)")
    c.font = SECTION_FONT
    r += 1
    for q in QUARTERS:
        top = var["exceptions_summary"].get(q, [])
        if not top:
            continue
        _cell(ws, r, 1, q, bold=True)
        r += 1
        for row in top:
            _cell(ws, r, 2, row["side"])
            _cell(ws, r, 3, f'{row["name"]} [{row["code"]}]')
            _cell(ws, r, 4, row["actual"], AMT_FMT)
            _cell(ws, r, 5, row["projected"], AMT_FMT)
            _cell(ws, r, 6, row["var_abs"], AMT_FMT)
            _cell(ws, r, 9, row["remarks"])
            r += 1


def _sheet_qoq(ws, var):
    _header(ws, 1, ["Side", "Category", "Quarter Pair", "Actuals Δ (INR)",
                    "Projections Δ (INR)"], [9, 42, 13, 18, 18])
    r = 2
    for row in var["qoq"]:
        _cell(ws, r, 1, row["side"])
        _cell(ws, r, 2, f'{row["name"]} [{row["code"]}]')
        _cell(ws, r, 3, row["quarter_pair"])
        _cell(ws, r, 4, row["actual_delta"], AMT_FMT)
        _cell(ws, r, 5, row["projected_delta"], AMT_FMT)
        r += 1


def _sheet_exceptions(ws, wf_exceptions, manual_conds):
    headers = ["Side", "Violation Type (reason code)", "Cycle", "Date", "Quarter",
               "Source Row", "Category", "Amount (INR)", "Commentary"]
    _header(ws, 1, headers, [8, 30, 10, 12, 9, 11, 32, 16, 70])
    r = 2
    for e in sorted(wf_exceptions, key=lambda x: (x["side"], x["violation_type"], x["date"])):
        _cell(ws, r, 1, e["side"], fill=FLAG_FILL)
        _cell(ws, r, 2, e["violation_type"], fill=FLAG_FILL)
        _cell(ws, r, 3, e["cycle"], fill=FLAG_FILL)
        _cell(ws, r, 4, e["date"].strftime("%d-%m-%Y"), fill=FLAG_FILL)
        _cell(ws, r, 5, e["quarter"], fill=FLAG_FILL)
        _cell(ws, r, 6, e["row_no"], fill=FLAG_FILL)
        _cell(ws, r, 7, e["category"], fill=FLAG_FILL)
        _cell(ws, r, 8, e["amount"], AMT_FMT, fill=FLAG_FILL)
        _cell(ws, r, 9, e["detail"], fill=FLAG_FILL)
        r += 1
    if not wf_exceptions:
        _cell(ws, r, 1, "No waterfall/condition violations detected.")
        r += 1
    r += 1
    c = ws.cell(row=r, column=1, value="Sanction conditions requiring manual verification (not machine-checkable):")
    c.font = SECTION_FONT
    r += 1
    for cond in manual_conds:
        _cell(ws, r, 1, cond.get("id", ""))
        _cell(ws, r, 2, "MANUAL_CHECK")
        _cell(ws, r, 9, cond.get("description", ""))
        r += 1


def _sheet_review_queue(ws, txns):
    headers = ["Queue", "Source Row", "Date", "Type", "Amount (INR)", "Narration",
               "Reason", "Analyst Category (fill in)", "Analyst Remarks (fill in)"]
    _header(ws, 1, headers, [22, 11, 12, 6, 16, 50, 42, 26, 32])
    r = 2
    fill_in = PatternFill("solid", fgColor="FFFF00")
    for t in txns:
        if not (t.needs_review or t.is_escalated_conflict):
            continue
        queue = ("ESCALATED — MANUAL OVERRIDE" if t.is_escalated_conflict
                 else "ANALYST REVIEW")
        _cell(ws, r, 1, queue, bold=t.is_escalated_conflict)
        _cell(ws, r, 2, t.row_no)
        _cell(ws, r, 3, t.txn_date.strftime("%d-%m-%Y"))
        _cell(ws, r, 4, t.txn_type)
        _cell(ws, r, 5, t.amount, AMT_FMT)
        _cell(ws, r, 6, t.narration or "(blank narration)")
        _cell(ws, r, 7, t.conflict_reason or t.review_reason)
        _cell(ws, r, 8, "", fill=fill_in)
        _cell(ws, r, 9, "", fill=fill_in)
        r += 1
    if r == 2:
        _cell(ws, r, 1, "Review queue is empty — all transactions classified.")
        r += 1
    r += 1
    legend = ws.cell(row=r, column=1,
                     value="Legend: yellow cells are for analyst input — enter the confirmed category code and "
                           "remarks, then re-run the agent with the reviewed file if reclassification is needed. "
                           "Example: Analyst Category = DEBT_SERV, Remarks = 'Confirmed with lender advice dated 05-08-2025'.")
    legend.font = Font(name=FONT, italic=True, size=9)


def _sheet_audit(ws, audit):
    _header(ws, 1, ["Timestamp", "Processing Step", "Detail"], [20, 24, 120])
    for r, e in enumerate(audit.entries, start=2):
        _cell(ws, r, 1, e["timestamp"])
        _cell(ws, r, 2, e["step"])
        _cell(ws, r, 3, e["detail"])
