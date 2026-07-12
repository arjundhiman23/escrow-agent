"""Build the TRA analysis from the Sanction Note (CAM) projected P&L.

Produces TRA_Analysis.xlsx:
  - Projected P&L (as extracted, FY26-FY35)
  - TRA cashflow view mapped to CATRA codes (projected inflows/outflows per FY,
    plus H1/H2 semi-annual split aligned to the biannual annuity cycle)
  - Reserve adequacy checks (DSRA / WCR / MMR) vs Agreement covenants,
    each marked COMPLIANT / UNDERBREACH / OVERBREACH on projected basis.
Returns the reserve-check rows so the Final Analysis can reuse the verdicts.
"""
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SEC_FILL = PatternFill("solid", fgColor="D9E2F3")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
AMB_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
NUM = "#,##0.00"

# income/opex line -> CATRA code
INFLOW_MAP = {"TPC Annuity": "ANNUITY", "O&M Receipts": "ANNUITY", "Interest on Annuity": "ANNUITY"}
OUTFLOW_MAP = {"Routine Maintenance": "OM_EXP", "Other O&M": "OM_EXP", "MMR Provisioning": "MMRA"}


def _h(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
        c.fill = HDR_FILL
        c.border = THIN
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def _row(ws, r, label, vals, bold=False, sec=False, indent=0):
    c = ws.cell(row=r, column=1, value=("    " * indent) + label)
    c.font = Font(name=ARIAL, size=10, bold=bold)
    c.border = THIN
    if sec:
        c.fill = SEC_FILL
    for j, v in enumerate(vals, 2):
        cell = ws.cell(row=r, column=j, value=v)
        cell.font = Font(name=ARIAL, size=10, bold=bold)
        cell.border = THIN
        cell.number_format = NUM
        if sec:
            cell.fill = SEC_FILL


def reserve_checks(extract):
    """Compute reserve adequacy on projected basis. Returns list of dicts.

    Statuses: COMPLIANT / MARGINAL (within proxy tolerance) / UNDERBREACH / N-A.
    """
    pnl, bs = extract["projected_pnl"], extract["projected_balance_sheet"]
    fye = pnl["fye"]
    interest, cmltd = pnl["interest"], bs["cmltd"]
    checks = []
    for i, fy in enumerate(fye):
        # CV2: DSRA = forthcoming 2 quarters debt service ~ (interest + CMLTD)/2 of the following year
        j = min(i + 1, len(fye) - 1)
        req_dsra = round((interest[j] + cmltd[j]) / 2, 2)
        have = bs["dsra_fund"][i]
        gap = round(have - req_dsra, 2)
        if i == 0:
            status = "N-A"
            note = "Construction/COD transition year — DSRA is an Operational Period requirement; one-time opening DSRA due from Promoter on/before COD"
        elif have + 1e-9 >= req_dsra:
            status, note = "COMPLIANT", "Fund meets 2-quarter proxy requirement"
        elif abs(gap) <= 0.05 * req_dsra:
            status = "MARGINAL"
            note = "Gap within 5% of proxy requirement — proxy uses (next-FY interest + CMLTD)/2; verify against actual forthcoming-2-quarter schedule"
        else:
            status, note = "UNDERBREACH", "Fund below 2-quarter debt service proxy requirement"
        checks.append({"covenant": "CV2 DSRA (2 quarters debt service)", "fy": fy,
                       "required": req_dsra, "provided": have, "gap": gap,
                       "status": status, "note": note})
    for i, fy in enumerate(fye):
        # CV1: WCR = 6 months interest servicing (per agreement wording)
        req_wcr = round(interest[i] / 2, 2)
        have = bs["working_capital_reserve"][i]
        checks.append({
            "covenant": "CV1 WCR (6 months interest)", "fy": fy,
            "required": req_wcr, "provided": have, "gap": round(have - req_wcr, 2),
            "status": "COMPLIANT" if have + 1e-9 >= req_wcr else "UNDERBREACH",
            "note": "Agreement 2.3(B)(i) sizes WCR at 6 months interest at COD; CAM base case carries a materially smaller WCR — interpretation to be confirmed with lender",
        })
    # CV3: MMR fund vs cumulative provisioning since last drawdown
    TOL = 0.05
    cum, prev_fund = 0.0, 0.0
    for i, fy in enumerate(fye):
        prov = pnl["opex"]["MMR Provisioning"][i]
        have = bs["mmr_fund"][i]
        if i and have < prev_fund - TOL:      # major-maintenance drawdown year: cycle resets
            cum = prov
            status, req = "N-A", have
            note = "Major-maintenance drawdown year — provisioning cycle resets; fund utilisation per 2.3(B)(k)"
            gap = 0.0
        else:
            cum = round(cum + prov, 2)
            req = cum
            gap = round(have - cum, 2)
            status = "COMPLIANT" if gap >= -TOL else "UNDERBREACH"
            note = "MMR fund tracks cumulative base-case provisioning since last drawdown (tolerance Rs 0.05 cr)"
        prev_fund = have
        checks.append({"covenant": "CV3 MMRA (per base case)", "fy": fy,
                       "required": req, "provided": have, "gap": gap,
                       "status": status, "note": note})
    return checks


def build_tra_xlsx(extract, out_path="output/TRA_Analysis.xlsx"):
    pnl, bs = extract["projected_pnl"], extract["projected_balance_sheet"]
    fye = pnl["fye"]
    wb = Workbook()

    # ---- Sheet 1: Projected P&L (source) ----
    ws = wb.active
    ws.title = "Projected P&L (CAM)"
    _h(ws, 1, ["Rs. crore"] + fye, [30] + [11] * len(fye))
    r = 2
    _row(ws, r, "Income", [None] * len(fye), bold=True, sec=True); r += 1
    for k, v in pnl["income"].items():
        _row(ws, r, k, v, indent=1); r += 1
    _row(ws, r, "Total Income", pnl["income_total"], bold=True); r += 1
    _row(ws, r, "Operating Expenses", [None] * len(fye), bold=True, sec=True); r += 1
    for k, v in pnl["opex"].items():
        _row(ws, r, k, v, indent=1); r += 1
    _row(ws, r, "Total Opex", pnl["opex_total"], bold=True); r += 1
    for label, key in [("EBITDA", "ebitda"), ("Interest", "interest"),
                       ("Depreciation & Amortisation", "depreciation"), ("PBT", "pbt"),
                       ("CSR", "csr"), ("Tax", "tax"), ("PAT", "pat")]:
        _row(ws, r, label, pnl[key], bold=label in ("EBITDA", "PBT", "PAT")); r += 1
    ws.freeze_panes = "B2"

    # ---- Sheet 2: TRA cashflow mapped to CATRA ----
    ws2 = wb.create_sheet("TRA Cashflow (CATRA-mapped)")
    _h(ws2, 1, ["Flow", "CATRA Code", "Line"] + fye, [9, 13, 26] + [11] * len(fye))
    r = 2
    for k, v in pnl["income"].items():
        _row2 = [ws2.cell(row=r, column=1, value="Inflow"), ws2.cell(row=r, column=2, value=INFLOW_MAP[k]),
                 ws2.cell(row=r, column=3, value=k)]
        for c in _row2: c.font = Font(name=ARIAL, size=10); c.border = THIN
        for j, val in enumerate(v, 4):
            cell = ws2.cell(row=r, column=j, value=val); cell.font = Font(name=ARIAL, size=10)
            cell.border = THIN; cell.number_format = NUM
        r += 1
    outrows = [("Routine Maintenance", pnl["opex"]["Routine Maintenance"]),
               ("Other O&M", pnl["opex"]["Other O&M"]),
               ("MMR Provisioning", pnl["opex"]["MMR Provisioning"]),
               ("Interest (Facility)", pnl["interest"]),
               ("Principal (CMLTD)", bs["cmltd"]),
               ("CSR / Statutory", pnl["csr"])]
    OUT_CODE = {"Routine Maintenance": "OM_EXP", "Other O&M": "OM_EXP", "MMR Provisioning": "MMRA",
                "Interest (Facility)": "DEBT_SERV", "Principal (CMLTD)": "DEBT_SERV", "CSR / Statutory": "TAXES"}
    for k, v in outrows:
        cells = [ws2.cell(row=r, column=1, value="Outflow"), ws2.cell(row=r, column=2, value=OUT_CODE[k]),
                 ws2.cell(row=r, column=3, value=k)]
        for c in cells: c.font = Font(name=ARIAL, size=10); c.border = THIN
        for j, val in enumerate(v, 4):
            cell = ws2.cell(row=r, column=j, value=val); cell.font = Font(name=ARIAL, size=10)
            cell.border = THIN; cell.number_format = NUM
        r += 1
    # net surplus row (formula)
    lab = ws2.cell(row=r, column=3, value="Net surplus before reserves/sweep")
    lab.font = Font(name=ARIAL, size=10, bold=True); lab.border = THIN
    for j in range(4, 4 + len(fye)):
        col = get_column_letter(j)
        cell = ws2.cell(row=r, column=j, value=f"=SUM({col}2:{col}4)-SUM({col}5:{col}{r-1})")
        cell.font = Font(name=ARIAL, size=10, bold=True); cell.border = THIN; cell.number_format = NUM
    ws2.freeze_panes = "D2"

    # ---- Sheet 3: Semi-annual profile ----
    ws3 = wb.create_sheet("Semi-Annual Profile")
    _h(ws3, 1, ["FY", "Half", "Projected Annuity-linked Inflow", "Projected Debt Service", "Basis"],
       [8, 8, 26, 22, 60])
    r = 2
    for i, fy in enumerate(fye):
        inflow = pnl["income_total"][i]
        ds = pnl["interest"][i] + bs["cmltd"][i]
        for half in ("H1", "H2"):
            vals = [fy, half, round(inflow / 2, 2), round(ds / 2, 2),
                    "Annuity payable biannually (30 instalments from 180th day of COD); 50/50 split of FY projection pending instalment-wise schedule"]
            for j, v in enumerate(vals, 1):
                c = ws3.cell(row=r, column=j, value=v)
                c.font = Font(name=ARIAL, size=10); c.border = THIN
                if j in (3, 4): c.number_format = NUM
            r += 1
    ws3.freeze_panes = "A2"

    # ---- Sheet 4: Reserve adequacy ----
    checks = reserve_checks(extract)
    ws4 = wb.create_sheet("Reserve Adequacy")
    _h(ws4, 1, ["Covenant", "FY", "Required (Rs cr)", "Provided (Rs cr)", "Gap", "Status", "Note"],
       [34, 8, 15, 15, 12, 15, 60])
    for i, ck in enumerate(checks):
        r = i + 2
        vals = [ck["covenant"], ck["fy"], ck["required"], ck["provided"], ck["gap"], ck["status"], ck["note"]]
        for j, v in enumerate(vals, 1):
            c = ws4.cell(row=r, column=j, value=v)
            c.font = Font(name=ARIAL, size=10); c.border = THIN
            c.alignment = Alignment(vertical="top", wrap_text=(j == 7))
            if j in (3, 4, 5): c.number_format = NUM
        st = ws4.cell(row=r, column=6)
        st.fill = (GRN_FILL if ck["status"] == "COMPLIANT"
                   else AMB_FILL if ck["status"] in ("MARGINAL", "N-A")
                   else RED_FILL)
        st.font = Font(name=ARIAL, size=10, bold=True)
    ws4.freeze_panes = "A2"

    wb.save(out_path)
    return out_path, checks
